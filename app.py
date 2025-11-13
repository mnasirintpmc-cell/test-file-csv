import streamlit as st
import pandas as pd
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.data_validation import DataValidation
import tempfile
import os

# Configure the page
st.set_page_config(
    page_title="Main Seal Test Manager",
    page_icon="⚙️",
    layout="wide"
)

def create_styled_excel_template():
    """Create a beautifully formatted Excel template with dropdowns and styling"""
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet and create our formatted sheets
    wb.remove(wb.active)
    
    # Create Instructions sheet
    ws_instructions = wb.create_sheet("INSTRUCTIONS")
    
    # Instruction content
    instructions = [
        ["MAIN SEAL TEST SEQUENCE TEMPLATE - INSTRUCTIONS"],
        [""],
        ["HOW TO USE THIS TEMPLATE:"],
        ["1. Fill in the TEST_SEQUENCE sheet with your test parameters"],
        ["2. Use dropdown menus for fields with limited options (Yes/No, Mode, Gas Type)"],
        ["3. Required fields: Step, Speed, Cell Pressure, Duration, Temperature"],
        ["4. Save your completed file and upload back to the web app"],
        ["5. Download the generated CSV for your control system"],
        [""],
        ["FIELD DESCRIPTIONS:"],
        ["Step: Sequential test step number (1, 2, 3...)"],
        ["Speed_RPM: Rotational speed (0 = stationary, 3600 = max speed)"],
        ["Cell_Pressure_bar: Main chamber pressure (0.1 - 100 bar)"],
        ["Interface_Pressure_bar: Interface pressure (0-40 bar)"],
        ["BP_Drive_End_bar: Back pressure drive end (0-7 bar)"],
        ["BP_Non_Drive_End_bar: Back pressure non-drive end (0-7 bar)"],
        ["Gas_Injection_bar: Gas injection pressure (0-5 bar)"],
        ["Duration_s: Step duration in seconds (1-300)"],
        ["Auto_Proceed: Automatic step progression (Yes/No)"],
        ["Temperature_C: Test temperature (30-155°C)"],
        ["Gas_Type: Test gas type (Air/N2/He)"],
        ["Test_Mode: Operating mode (Mode 1/Mode 2)"],
        ["Measurement: Take measurements (Yes/No)"],
        ["Torque_Check: Perform torque check (Yes/No)"],
        ["Notes: Additional comments or observations"]
    ]
    
    # Add instructions to sheet
    for row_idx, instruction in enumerate(instructions, 1):
        ws_instructions.cell(row=row_idx, column=1, value=instruction[0])
    
    # Style instructions sheet
    title_font = Font(size=14, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    ws_instructions.cell(1, 1).font = title_font
    ws_instructions.cell(1, 1).fill = title_fill
    
    for row in range(3, 25):
        if ws_instructions.cell(row, 1).value and ":" in ws_instructions.cell(row, 1).value:
            ws_instructions.cell(row, 1).font = header_font
            ws_instructions.cell(row, 1).fill = header_fill
    
    # Create main data sheet
    ws_data = wb.create_sheet("TEST_SEQUENCE")
    
    # Define headers with descriptions
    headers = [
        "Step", "Speed_RPM", "Cell_Pressure_bar", "Interface_Pressure_bar",
        "BP_Drive_End_bar", "BP_Non_Drive_End_bar", "Gas_Injection_bar",
        "Duration_s", "Auto_Proceed", "Temperature_C", "Gas_Type", 
        "Test_Mode", "Measurement", "Torque_Check", "Notes"
    ]
    
    # Add headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add sample data
    sample_data = [
        [1, 0, 0.21, 0, 0, 0, 0, 2, "No", 30, "Air", "Mode 1", "Yes", "No", "Initial low pressure test"],
        [2, 0, 0.4, 0, 0, 0, 0, 2, "No", 30, "Air", "Mode 1", "Yes", "No", "Pressure increase"],
        [3, 0, 1.0, 0, 0, 0, 0, 2, "No", 30, "Air", "Mode 1", "Yes", "No", "Medium pressure check"],
        [4, 3600, 10.0, 0, 1.0, 1.0, 0, 10, "Yes", 155, "Air", "Mode 1", "Yes", "No", "High speed operation"],
        [5, 0, 0, 0, 0, 0, 0, 1, "No", 155, "Air", "Mode 1", "No", "No", "System cool down"]
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_data.cell(row=row_idx, column=col_idx, value=value)
    
    # Apply styling to data
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply borders and alignment to all cells
    for row in ws_data.iter_rows(min_row=1, max_row=len(sample_data)+1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
            if cell.row == 1:  # Header row
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data validation for dropdowns
    # Auto_Proceed dropdown
    auto_proceed_dv = DataValidation(
        type="list", 
        formula1='"Yes,No"',
        allow_blank=False
    )
    ws_data.add_data_validation(auto_proceed_dv)
    auto_proceed_dv.add(f'I2:I100')
    
    # Gas_Type dropdown
    gas_type_dv = DataValidation(
        type="list",
        formula1='"Air,N2,He"',
        allow_blank=False
    )
    ws_data.add_data_validation(gas_type_dv)
    gas_type_dv.add(f'K2:K100')
    
    # Test_Mode dropdown
    test_mode_dv = DataValidation(
        type="list",
        formula1='"Mode 1,Mode 2"',
        allow_blank=False
    )
    ws_data.add_data_validation(test_mode_dv)
    test_mode_dv.add(f'L2:L100')
    
    # Measurement dropdown
    measurement_dv = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False
    )
    ws_data.add_data_validation(measurement_dv)
    measurement_dv.add(f'M2:M100')
    
    # Torque_Check dropdown
    torque_dv = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False
    )
    ws_data.add_data_validation(torque_dv)
    torque_dv.add(f'N2:N100')
    
    # Set column widths
    column_widths = [8, 12, 18, 20, 18, 22, 18, 12, 12, 15, 12, 12, 12, 12, 30]
    for col_idx, width in enumerate(column_widths, 1):
        ws_data.column_dimensions[chr(64 + col_idx)].width = width
    
    # Add conditional formatting for important fields
    # Highlight speed changes
    for row in range(2, len(sample_data) + 2):
        speed_cell = ws_data.cell(row=row, column=2)  # Speed column
        if speed_cell.value and speed_cell.value > 0:
            speed_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        # Highlight auto-proceed steps
        auto_cell = ws_data.cell(row=row, column=9)  # Auto Proceed column
        if auto_cell.value == "Yes":
            auto_cell.fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
    
    return wb

def excel_to_machine_csv(excel_file):
    """Convert technician Excel to machine-readable CSV"""
    
    # Read Excel file (use TEST_SEQUENCE sheet)
    df = pd.read_excel(excel_file, sheet_name='TEST_SEQUENCE')
    
    # Remove empty rows
    df = df.dropna(subset=['Step']).reset_index(drop=True)
    
    # Convert back to machine format
    machine_mapping = {
        'Speed_RPM': 'TST_SpeedDem',
        'Cell_Pressure_bar': 'TST_CellPresDemand',
        'Interface_Pressure_bar': 'TST_InterPresDemand',
        'BP_Drive_End_bar': 'TST_InterBPDemand_DE',
        'BP_Non_Drive_End_bar': 'TST_InterBPDemand_NDE',
        'Gas_Injection_bar': 'TST_GasInjectionDemand',
        'Duration_s': 'TST_StepDuration',
        'Auto_Proceed': 'TST_APFlag',
        'Temperature_C': 'TST_TempDemand',
        'Gas_Type': 'TST_GasType',
        'Test_Mode': 'TST_TestMode',
        'Measurement': 'TST_MeasurementReq',
        'Torque_Check': 'TST_TorqueCheck'
    }
    
    # Rename columns
    df = df.rename(columns=machine_mapping)
    
    # Convert readable values back to machine codes
    df['TST_APFlag'] = df['TST_APFlag'].map({'Yes': 1, 'No': 0})
    df['TST_MeasurementReq'] = df['TST_MeasurementReq'].map({'Yes': 1, 'No': 0})
    df['TST_TorqueCheck'] = df['TST_TorqueCheck'].map({'Yes': 1, 'No': 0})
    df['TST_TestMode'] = df['TST_TestMode'].map({'Mode 1': 1, 'Mode 2': 2})
    
    # Remove Step and Notes columns for machine CSV
    machine_df = df.drop(['Step', 'Notes'], axis=1, errors='ignore')
    
    return machine_df

def machine_csv_to_excel(csv_file):
    """Convert machine CSV to formatted technician Excel"""
    
    # Read machine CSV
    df = pd.read_csv(csv_file, delimiter=';')
    
    # Convert to technician format
    technician_mapping = {
        'TST_SpeedDem': 'Speed_RPM',
        'TST_CellPresDemand': 'Cell_Pressure_bar',
        'TST_InterPresDemand': 'Interface_Pressure_bar',
        'TST_InterBPDemand_DE': 'BP_Drive_End_bar',
        'TST_InterBPDemand_NDE': 'BP_Non_Drive_End_bar',
        'TST_GasInjectionDemand': 'Gas_Injection_bar',
        'TST_StepDuration': 'Duration_s',
        'TST_APFlag': 'Auto_Proceed',
        'TST_TempDemand': 'Temperature_C',
        'TST_GasType': 'Gas_Type',
        'TST_TestMode': 'Test_Mode',
        'TST_MeasurementReq': 'Measurement',
        'TST_TorqueCheck': 'Torque_Check'
    }
    
    # Rename columns
    df = df.rename(columns=technician_mapping)
    
    # Convert machine codes to readable values
    df['Auto_Proceed'] = df['Auto_Proceed'].map({0: 'No', 1: 'Yes'})
    df['Measurement'] = df['Measurement'].map({0: 'No', 1: 'Yes'})
    df['Torque_Check'] = df['Torque_Check'].map({0: 'No'})
    df['Test_Mode'] = df['Test_Mode'].map({1: 'Mode 1', 2: 'Mode 2'})
    
    # Add Step numbers and Notes column
    df.insert(0, 'Step', range(1, len(df) + 1))
    df['Notes'] = ''  # Empty notes column for technicians
    
    return df

def main():
    st.title("⚙️ Main Seal Test Sequence Manager")
    st.markdown("### Professional Template System for Technicians")
    
    # Sidebar
    st.sidebar.title("🔧 Operations")
    operation = st.sidebar.radio(
        "Choose Operation:",
        ["📥 Download Template", "🔄 Excel to Machine CSV", "📤 Machine CSV to Excel", "👀 View Current Test"]
    )
    
    if operation == "📥 Download Template":
        st.header("Download Professional Excel Template")
        
        st.info("""
        **🎯 Features of this template:**
        - 🎨 **Professional formatting** with colored headers and borders
        - 📋 **Dropdown menus** for standardized inputs
        - 📝 **Built-in instructions** and examples
        - 🎯 **Data validation** to prevent errors
        - 💡 **Conditional formatting** for important values
        """)
        
        # Create and download template
        wb = create_styled_excel_template()
        
        # Save to temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name
        
        with open(tmp_path, 'rb') as f:
            st.download_button(
                label="📥 Download Professional Template (.xlsx)",
                data=f,
                file_name=f"main_seal_template_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        # Clean up
        os.unlink(tmp_path)
        
        st.subheader("🎨 Template Features Preview")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("""
            **📋 Dropdown Selections:**
            - Auto Proceed: Yes/No
            - Gas Type: Air/N2/He  
            - Test Mode: Mode 1/Mode 2
            - Measurement: Yes/No
            - Torque Check: Yes/No
            """)
        
        with col2:
            st.markdown("""
            **🎯 Visual Highlights:**
            - Blue headers with white text
            - Clear cell borders
            - Yellow highlight for speed > 0
            - Green highlight for auto-proceed steps
            - Centered text alignment
            """)
        
        st.markdown("""
        **📁 Template Structure:**
        - **INSTRUCTIONS** sheet: Detailed guide and field descriptions
        - **TEST_SEQUENCE** sheet: Main data entry with formatting
        - **Sample data** included for reference
        """)
    
    elif operation == "🔄 Excel to Machine CSV":
        st.header("Convert Excel to Machine CSV")
        
        uploaded_file = st.file_uploader("Upload your completed Excel template", type=['xlsx'])
        
        if uploaded_file:
            try:
                # Convert to machine format
                machine_df = excel_to_machine_csv(uploaded_file)
                
                st.success(f"✅ Successfully converted {len(machine_df)} test steps!")
                
                # Preview
                st.subheader("Machine CSV Preview")
                st.dataframe(machine_df, use_container_width=True)
                
                # Show conversion summary
                col1, col2 = st.columns(2)
                with col1:
                    st.metric("Total Steps", len(machine_df))
                with col2:
                    st.metric("Auto Proceed Steps", 
                             len(machine_df[machine_df['TST_APFlag'] == 1]))
                
                # Download machine CSV
                csv_data = machine_df.to_csv(index=False, sep=';')
                st.download_button(
                    label="📥 Download Machine CSV",
                    data=csv_data,
                    file_name=f"machine_sequence_{datetime.now().strftime('%Y%m%d')}.csv",
                    mime="text/csv"
                )
                
            except Exception as e:
                st.error(f"❌ Error converting file: {str(e)}")
                st.info("Make sure you're using the provided template format.")
    
    elif operation == "📤 Machine CSV to Excel":
        st.header("Convert Machine CSV to Excel")
        
        uploaded_file = st.file_uploader("Upload machine CSV file", type=['csv'])
        
        if uploaded_file:
            try:
                # Convert to technician format
                technician_df = machine_csv_to_excel(uploaded_file)
                
                st.success(f"✅ Successfully converted {len(technician_df)} test steps!")
                
                # Preview
                st.subheader("Converted Data Preview")
                st.dataframe(technician_df, use_container_width=True)
                
                # Create formatted Excel for download
                wb = Workbook()
                ws = wb.active
                ws.title = "TEST_SEQUENCE"
                
                # Add headers with basic formatting
                for col_idx, column in enumerate(technician_df.columns, 1):
                    cell = ws.cell(row=1, column=col_idx, value=column)
                    cell.font = Font(bold=True)
                
                # Add data
                for row_idx, row in technician_df.iterrows():
                    for col_idx, value in enumerate(row, 1):
                        ws.cell(row=row_idx+2, column=col_idx, value=value)
                
                # Save to temporary file
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                    wb.save(tmp.name)
                    tmp_path = tmp.name
                
                with open(tmp_path, 'rb') as f:
                    st.download_button(
                        label="📥 Download Technician Excel",
                        data=f,
                        file_name=f"technician_sequence_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                
                # Clean up
                os.unlink(tmp_path)
                
            except Exception as e:
                st.error(f"❌ Error converting file: {str(e)}")
    
    elif operation == "👀 View Current Test":
        st.header("Current Test Sequence")
        
        # Try to load the existing CSV
        try:
            df = pd.read_csv('MainSealSet2.csv', delimiter=';')
            technician_df = machine_csv_to_excel('MainSealSet2.csv')
            
            st.success(f"✅ Loaded existing test sequence with {len(df)} steps")
            
            # Display metrics
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Total Steps", len(df))
            with col2:
                st.metric("Temperature Range", f"{df['TST_TempDemand'].min()}°C - {df['TST_TempDemand'].max()}°C")
            with col3:
                st.metric("Max Speed", f"{df['TST_SpeedDem'].max()} RPM")
            with col4:
                st.metric("Auto Proceed Steps", len(df[df['TST_APFlag'] == 1]))
            
            # Display in technician-friendly format
            st.subheader("Current Test Sequence")
            st.dataframe(technician_df, use_container_width=True, height=500)
            
        except Exception as e:
            st.error(f"❌ Could not load current test sequence: {str(e)}")
            st.info("Upload a file using the conversion tools above to get started.")

if __name__ == "__main__":
    main()
